import openpyxl;
wb_obj = openpyxl.load_workbook(r"bayes.xlsx")
sheet = wb_obj.active
n = sheet.max_row
m=sheet.max_column

total=0
for i in range(1,n+1):
    if(sheet.cell(row=i,column=m).value=="Play"):
        total+=1
totalNoPlay=(n-1)-total

list_play=[]
list_noPlay=[]
for i in range(2,m):
    b=sheet.cell(row=1,column=i).value
    a = sheet.cell(row=1,column=i).value
    sunny=0
    non_sunny=0
    for j in range(2,n+1):
        if(sheet.cell(row=j,column=i).value==a):
            if(sheet.cell(row=j,column=m).value=="Outlook"):
                sunny+=1
            else:
                non_sunny+=1
    list_play.append(outlook/total)
    list_noPlay.append(non_sunny/totalNoPlay)


prob_play=1
prob_non_play=1
for i in list_noPlay:
    prob_play=prob_play*i
for i in list_noPlay:
    prob_non_play=prob_non_play*i
prob_play=prob_play*(total/(n-1))
prob_non_play=prob_non_play*(totalNoPlay/(n-1))


print("Probability for Sunny and Play is",prob_play)
print("Probability for Sunny and NoPlay is",prob_non_play)